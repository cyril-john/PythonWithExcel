import openpyxl as xl
from openpyxl.cell import WriteOnlyCell

wb = xl.load_workbook('Book1.xlsx')
ws = wb.create_sheet()
sheet = wb['Sheet1']
cell = sheet.cell(2,2)
print(cell.value)
print(cell.comment.text)
cell1 = WriteOnlyCell(sheet,value="changeIn Name")
my_list = [1, 2,3,3,3,3]
ws.append(my_list)
wb.save('sample.xlsx')


